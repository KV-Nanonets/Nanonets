
from docx import Document
import openpyxl

def convert_word_tables_to_excel(word_file_path, excel_file_path):
    # Load the Word document
    doc = Document(word_file_path)

    # Create a new Excel workbook
    wb = openpyxl.Workbook()

    # Loop through each table in the Word document
    for idx, table in enumerate(doc.tables):
        # Create a new sheet for each table
        ws = wb.create_sheet(title=f"Table {idx + 1}")

        # Loop through each row in the table
        for row_idx, row in enumerate(table.rows):
            # Loop through each cell in the row
            for col_idx, cell in enumerate(row.cells):
                # Write the cell value to the corresponding Excel cell
                ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell.text)

    # Remove the default sheet created by openpyxl
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Save the Excel workbook
    wb.save(excel_file_path)
    print(f"Converted Word tables to Excel: {excel_file_path}")

def main():
    # Example usage
    word_file_path = 'your_word_document.docx'  # Replace with your Word document path
    excel_file_path = 'output_excel_file.xlsx'  # Replace with desired Excel file path

    convert_word_tables_to_excel(word_file_path, excel_file_path)

if __name__ == "__main__":
    main()
